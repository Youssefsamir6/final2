"""
Daily Attendance Report Generator
Reads Attendance.csv and produces a formatted Excel report.
Run standalone or call generate_daily_report() from your main script.
"""

import os
import csv
from datetime import datetime, date
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


ATTENDANCE_FILE = "Attendance.csv"
REPORTS_FOLDER  = "Reports"

# ── Colour palette ────────────────────────────────────────
C_DARK_BG   = "1F3864"   # deep navy  – header bg
C_MID_BG    = "2E75B6"   # medium blue – sub-header bg
C_ACCENT    = "00B0F0"   # sky blue   – highlights
C_PRESENT   = "E2EFDA"   # soft green – present rows
C_ABSENT    = "FCE4D6"   # soft red   – absent rows
C_WHITE     = "FFFFFF"
C_LIGHT_ROW = "F2F7FB"   # alternating light row
C_BORDER    = "BDD7EE"   # light blue border


# ── Style helpers ─────────────────────────────────────────
def _border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _style(cell, bold=False, fg=None, font_color="000000",
           font_size=11, align="left", border=True, italic=False, wrap=False):
    cell.font      = _font(bold=bold, color=font_color, size=font_size, italic=italic)
    cell.alignment = _align(h=align, wrap=wrap)
    if fg:
        cell.fill = _fill(fg)
    if border:
        cell.border = _border()

def _merge_header(ws, row, col_start, col_end, text, fg, font_color=C_WHITE,
                  font_size=12, bold=True):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    _style(cell, bold=bold, fg=fg, font_color=font_color,
           font_size=font_size, align="center")


# ── Data loading ──────────────────────────────────────────
def load_attendance(filepath=ATTENDANCE_FILE):
    records = []
    if not os.path.isfile(filepath):
        print(f"[WARN] {filepath} not found – generating empty report.")
        return records
    with open(filepath, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                records.append({
                    "name": row["Name"].strip(),
                    "date": row["Date"].strip(),
                    "time": row["Time"].strip(),
                })
            except KeyError:
                continue
    return records


# ── Sheet 1: Today's Attendance ───────────────────────────
def build_today_sheet(wb, records, target_date: str, roster: list):
    ws = wb.active
    ws.title = "Today"
    ws.sheet_view.showGridLines = False

    today_records = {r["name"]: r["time"] for r in records if r["date"] == target_date}

    # ---- Title block ----
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 16

    _merge_header(ws, 1, 1, 6,
                  f"  Attendance Report  –  {target_date}",
                  C_DARK_BG, font_size=16)

    generated = datetime.now().strftime("%d %b %Y  %H:%M")
    _merge_header(ws, 2, 1, 6,
                  f"Generated: {generated}",
                  C_MID_BG, font_size=10, bold=False)

    ws.row_dimensions[3].height = 8  # spacer

    # ---- Summary cards (row 4) ----
    ws.row_dimensions[4].height = 28
    total   = len(roster) if roster else len(today_records)
    present = len(today_records)
    absent  = max(0, total - present)
    rate    = f"{round(present / total * 100)}%" if total else "N/A"

    summaries = [
        ("Total Enrolled", str(total),   C_MID_BG,  C_WHITE),
        ("Present Today",  str(present), "375623",  C_WHITE),
        ("Absent Today",   str(absent),  "C00000",  C_WHITE),
        ("Attendance Rate",rate,          C_ACCENT,  "000000"),
    ]
    for i, (label, val, bg, fc) in enumerate(summaries):
        col = i * 2 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        c = ws.cell(row=4, column=col, value=f"{label}: {val}")
        _style(c, bold=True, fg=bg, font_color=fc, font_size=11, align="center")

    ws.row_dimensions[5].height = 8  # spacer

    # ---- Column headers ----
    ws.row_dimensions[6].height = 24
    headers = ["#", "Name", "Status", "Check-In Time", "Day", "Note"]
    col_widths = [5, 24, 14, 16, 12, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=6, column=col, value=h)
        _style(c, bold=True, fg=C_DARK_BG, font_color=C_WHITE, align="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    # ---- Data rows ----
    display_list = roster if roster else sorted(today_records.keys())
    if not display_list:
        display_list = sorted(today_records.keys())

    for idx, name in enumerate(display_list, 1):
        row = idx + 6
        ws.row_dimensions[row].height = 20
        is_present = name in today_records
        bg = C_PRESENT if is_present else C_ABSENT
        alt = C_LIGHT_ROW if idx % 2 == 0 else C_WHITE

        check_time = today_records.get(name, "—")
        status     = "Present" if is_present else "Absent"
        try:
            day_name = datetime.strptime(target_date, "%Y-%m-%d").strftime("%A")
        except ValueError:
            day_name = ""

        row_data = [idx, name, status, check_time, day_name, ""]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            row_bg = bg if col == 3 else alt
            status_bold = col == 3
            _style(c, bold=status_bold, fg=row_bg, align="center" if col != 2 else "left")

    # ---- Footer ----
    footer_row = len(display_list) + 8
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=6)
    c = ws.cell(row=footer_row, column=1,
                value="Face Recognition Attendance System  |  Auto-generated report")
    _style(c, italic=True, fg=C_DARK_BG, font_color="A0B4CC",
           font_size=9, align="center", border=False)


# ── Sheet 2: Full History ─────────────────────────────────
def build_history_sheet(wb, records):
    ws = wb.create_sheet("Full History")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 32
    _merge_header(ws, 1, 1, 4, "  Complete Attendance History", C_DARK_BG, font_size=14)
    ws.row_dimensions[2].height = 8

    ws.row_dimensions[3].height = 22
    for col, (h, w) in enumerate(
        [("Name", 24), ("Date", 14), ("Time", 14), ("Day", 14)], 1
    ):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, bold=True, fg=C_MID_BG, font_color=C_WHITE, align="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    for idx, rec in enumerate(sorted(records, key=lambda r: (r["date"], r["time"])), 1):
        row = idx + 3
        ws.row_dimensions[row].height = 18
        bg = C_LIGHT_ROW if idx % 2 == 0 else C_WHITE
        try:
            day_name = datetime.strptime(rec["date"], "%Y-%m-%d").strftime("%A")
        except ValueError:
            day_name = ""
        for col, val in enumerate([rec["name"], rec["date"], rec["time"], day_name], 1):
            c = ws.cell(row=row, column=col, value=val)
            _style(c, fg=bg, align="center" if col != 1 else "left")


# ── Sheet 3: Summary by Person ────────────────────────────
def build_summary_sheet(wb, records, roster: list):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 32
    _merge_header(ws, 1, 1, 5, "  Attendance Summary by Person", C_DARK_BG, font_size=14)
    ws.row_dimensions[2].height = 8

    all_dates = sorted({r["date"] for r in records})
    total_days = len(all_dates)

    by_person = defaultdict(set)
    for rec in records:
        by_person[rec["name"]].add(rec["date"])

    ws.row_dimensions[3].height = 22
    headers = ["Name", "Days Present", "Days Absent", "Total Days", "Attendance %"]
    widths  = [24, 14, 14, 12, 16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, bold=True, fg=C_MID_BG, font_color=C_WHITE, align="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    names = roster if roster else sorted(by_person.keys())
    data_start_row = 4

    for idx, name in enumerate(names, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 20
        bg = C_LIGHT_ROW if idx % 2 == 0 else C_WHITE
        days_present = len(by_person.get(name, set()))
        days_absent  = total_days - days_present
        pct_formula  = f"=IF(D{row}=0,0,B{row}/D{row})"

        row_data = [name, days_present, days_absent, total_days, pct_formula]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style(c, fg=bg, align="center" if col != 1 else "left")
            if col == 5:
                c.number_format = "0%"
                pct = days_present / total_days if total_days else 0
                c.fill = _fill(C_PRESENT if pct >= 0.75 else C_ABSENT)

    # Totals row
    totals_row = len(names) + 4
    ws.row_dimensions[totals_row].height = 22
    ws.cell(row=totals_row, column=1, value="TOTAL / AVERAGE")
    _style(ws.cell(row=totals_row, column=1),
           bold=True, fg=C_DARK_BG, font_color=C_WHITE, align="left")

    for col, formula in [
        (2, f"=SUM(B{data_start_row}:B{totals_row-1})"),
        (3, f"=SUM(C{data_start_row}:C{totals_row-1})"),
        (4, f"={total_days}"),
        (5, f"=AVERAGE(E{data_start_row}:E{totals_row-1})"),
    ]:
        c = ws.cell(row=totals_row, column=col, value=formula)
        _style(c, bold=True, fg=C_DARK_BG, font_color=C_WHITE, align="center")
        if col == 5:
            c.number_format = "0%"

    return ws, names, by_person, all_dates, data_start_row, len(names)


# ── Sheet 4: Daily Trend Chart ────────────────────────────
def build_chart_sheet(wb, records):
    ws = wb.create_sheet("Daily Trend")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 32
    _merge_header(ws, 1, 1, 3, "  Daily Attendance Trend", C_DARK_BG, font_size=14)
    ws.row_dimensions[2].height = 8

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    for col, h in enumerate(["Date", "Present", "Day"], 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, bold=True, fg=C_MID_BG, font_color=C_WHITE, align="center")

    by_date = defaultdict(int)
    for rec in records:
        by_date[rec["date"]] += 1

    dates_sorted = sorted(by_date.keys())
    for idx, d in enumerate(dates_sorted, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 18
        bg = C_LIGHT_ROW if idx % 2 == 0 else C_WHITE
        try:
            day_name = datetime.strptime(d, "%Y-%m-%d").strftime("%a")
        except ValueError:
            day_name = ""
        for col, val in enumerate([d, by_date[d], day_name], 1):
            c = ws.cell(row=row, column=col, value=val)
            _style(c, fg=bg, align="center")

    # Bar chart
    if len(dates_sorted) >= 2:
        chart = BarChart()
        chart.type    = "col"
        chart.grouping = "clustered"
        chart.title   = "Daily Attendance Count"
        chart.y_axis.title = "People Present"
        chart.x_axis.title = "Date"
        chart.style  = 10
        chart.width  = 22
        chart.height = 14

        data_ref = Reference(ws,
                             min_col=2, max_col=2,
                             min_row=3,
                             max_row=3 + len(dates_sorted))
        cats_ref = Reference(ws,
                             min_col=1, max_col=1,
                             min_row=4,
                             max_row=3 + len(dates_sorted))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = C_MID_BG

        ws.add_chart(chart, "E3")


# ── Master builder ────────────────────────────────────────
def generate_daily_report(
    attendance_file: str = ATTENDANCE_FILE,
    target_date:     str = None,
    roster:          list = None,
    output_folder:   str = REPORTS_FOLDER,
) -> str:
    """
    Generate a formatted Excel attendance report.

    Parameters
    ----------
    attendance_file : path to Attendance.csv
    target_date     : "YYYY-MM-DD" — defaults to today
    roster          : list of all enrolled names (for absence tracking)
    output_folder   : folder where the .xlsx is saved

    Returns
    -------
    Path to the saved .xlsx file.
    """
    if target_date is None:
        target_date = date.today().strftime("%Y-%m-%d")

    records = load_attendance(attendance_file)
    roster  = roster or []

    os.makedirs(output_folder, exist_ok=True)
    filename    = f"Attendance_Report_{target_date}.xlsx"
    output_path = os.path.join(output_folder, filename)

    wb = Workbook()

    build_today_sheet(wb, records, target_date, roster)
    build_history_sheet(wb, records)
    build_summary_sheet(wb, records, roster)
    build_chart_sheet(wb, records)

    # Tab colours
    wb["Today"].sheet_properties.tabColor        = C_MID_BG
    wb["Full History"].sheet_properties.tabColor = "4472C4"
    wb["Summary"].sheet_properties.tabColor      = "375623"
    wb["Daily Trend"].sheet_properties.tabColor  = C_ACCENT

    wb.save(output_path)
    print(f"[REPORT] Saved → {output_path}")
    return output_path


# ── Run standalone ────────────────────────────────────────
if __name__ == "__main__":
    ROSTER = [
        "Ahmed Hassan",
        "Sara Mohamed",
        "Mostafa Ali",
        "Nour Ibrahim",
        "Youssef Tarek",
    ]
    path = generate_daily_report(
        attendance_file=ATTENDANCE_FILE,
        roster=ROSTER,
    )
    print(f"Done: {path}")
